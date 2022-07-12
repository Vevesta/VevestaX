import json
import requests
import platform
import seaborn as sns
from datetime import datetime
from inspect import getmembers,stack,signature
from os.path import basename,expanduser,join,exists,isfile,isdir,islink
from os import mkdir,listdir,unlink
from random import randint
from re import sub
from shutil import copy,rmtree
from statistics import median,mode
from sys import argv
from uuid import getnode
from pathlib import Path
from github import Github,GithubException
from ipynbname import name
from matplotlib import pyplot as plt
from numpy import quantile,isinf
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from pandas import DataFrame as pandasDataframe
from pandas import ExcelWriter,read_excel,concat,Series
from pyspark.sql.dataframe import DataFrame as pysparkDataframe
from pyspark.ml.feature import VectorAssembler
from pyspark.ml.stat import Correlation
from pyspark.sql import SparkSession
from pyspark.sql.types import DoubleType
from scipy.stats import kurtosis
from scipy.stats import skew
from vevestaX import __version__
from img2pdf import mm_to_pt,get_layout_fun,convert

def test():
    return 'Test Executed Successfully'


class Experiment(object):
    def __init__(self, speedUp=False):
        self.__dataSourcing = None
        self.__featureEngineering = None
        self.__data = None
        self.__correlation = None

        self.__primitiveDataTypes = [int, str, float, bool]
        self.__startlocals = None
        self.__variables = {}
        self.__filename = self.get_filename()
        self.__sampleSize = 0
        self.__Y=None
        self.speedUp = speedUp

    def get_filename(self):
        try:
            try:
                filename = name() + '.ipynb'

            except:
                try:
                    filePath = dict(getmembers(stack()[2][0]))['f_locals']['__file__']
                except:
                    filePath = argv[0]
                filename = basename(filePath)
        except:
            filename = None
        return filename

    def __find_access_token(self):
        token = ""
        file_name = 'access_token.txt'

        directory = ".vevesta"
        parent_dir = expanduser("~")
        home_folder_path = join(parent_dir, directory)
        home_folder_file_path = join(home_folder_path, file_name)

        sibling_file_path = file_name

        if exists(sibling_file_path):
            token = self.__read_file(sibling_file_path)
            try:
                mkdir(home_folder_path)
            except FileExistsError:
                pass
            copy(sibling_file_path, home_folder_path)

        elif exists(home_folder_file_path):
            token = self.__read_file(home_folder_file_path)

        return token

    def __is_git_token_valid(self, token):
        try:
            Github(token).get_user().name
            return True
        except GithubException:
            return False

    def __read_file(self, path):
        f = open(path, 'r')
        data = f.read()
        f.close()
        return data

    def __write_file(self, path, data):
        f = open(path, "w+")
        f.write(data)
        f.close()

    def __find_git_token(self, is_v_commit, backend_url=None, access_token=None):
        file_name = 'git_token.txt'
        git_token = ''

        directory = ".vevesta"
        parent_dir = expanduser("~")
        home_folder_path = join(parent_dir, directory)

        sibling_file_path = file_name
        home_folder_file_path = join(home_folder_path, file_name)

        if exists(sibling_file_path) and self.__is_git_token_valid(
                self.__read_file(sibling_file_path)):
            git_token = self.__read_file(sibling_file_path)
            try:
                mkdir(home_folder_path)
            except FileExistsError:
                pass
            copy(sibling_file_path, home_folder_path)
            try:
                headers_for_set_token = {
                    'Authorization': 'Bearer ' + access_token,
                    'Access-Control-Allow-Origin': '*',
                    'Accept': '*/*',
                    'Content-Type': 'application/json'
                }
                payload = {
                    'gitToken': git_token
                }
                requests.post(url=backend_url + '/SetGitToken', headers=headers_for_set_token, data=json.dumps(payload))
            except:
                pass

        # it will be directly searched in the backend, if it’s called from within the V.commit function.
        elif is_v_commit:
            headers_for_git_token = {'Authorization': 'Bearer ' + access_token}
            response = requests.get(url=backend_url + '/GetGitToken', headers=headers_for_git_token)
            data = response.json()
            git_token = data['gitToken']
            if git_token != '' and self.__is_git_token_valid(git_token):
                try:
                    self.__write_file(sibling_file_path, git_token)
                    self.__write_file(home_folder_file_path, git_token)
                except:
                    pass
            else:
                raise Exception('Invalid Git Token')

        # The git token will be searched in .vevesta folder if commitToGit is called from the V.dump function
        elif not is_v_commit and exists(home_folder_file_path) and self.__is_git_token_valid(
                self.__read_file(home_folder_file_path)):
            git_token = self.__read_file(home_folder_file_path)

        return git_token

    def __fetch_project(self, backend_url, access_token, projectId):
        headers_for_project = {'Authorization': 'Bearer ' + access_token}
        params = {'projectId': projectId}
        response = requests.get(url=backend_url + '/Project', headers=headers_for_project, params=params)
        data = response.json()
        return data

    @property
    def dataSourcing(self):
        return self.__dataSourcing

    @dataSourcing.setter
    def dataSourcing(self, value):
        if isinstance(value, pandasDataframe):
            self.__dataSourcing = value.columns.tolist()
            self.__data = value
            self.__sampleSize = len(value)
            if self.speedUp == False:
                self.__correlation = value.corr(method='pearson')

        if isinstance(value, pysparkDataframe):
            self.__dataSourcing = value.columns
            self.__data = value
            self.__sampleSize = value.count()
            if self.speedUp == False:
                spark = SparkSession.builder.appName("vevesta").getOrCreate()
                columnNames = []
                columnNames = value.columns
                for i in range(len(value.columns)):
                    value = value.withColumn(columnNames[i], value[columnNames[i]].cast(DoubleType()))
                vectorCol = "corrfeatures"
                assembler = VectorAssembler(inputCols=value.columns, outputCol=vectorCol)
                df_vector = assembler.transform(value).select(vectorCol)
                matrix = Correlation.corr(df_vector, vectorCol).collect()[0][0]
                corrMatrix = matrix.toArray().tolist()
                dfCorr = spark.createDataFrame(corrMatrix, columnNames)
                self.__correlation = dfCorr

    @property
    def ds(self):  # its doing the same work as dataSourcing do
        return self.__dataSourcing

    @dataSourcing.setter
    def ds(self, value):
        self.dataSourcing = value

    @property
    def featureEngineering(self):
        return self.__featureEngineering

    @featureEngineering.setter
    def featureEngineering(self, value):
        if self.__dataSourcing is None:
            print("Data Sourcing step missed.")
            if isinstance(value, pandasDataframe):
                cols = value.columns
                self.__featureEngineering = cols

            if isinstance(value, pysparkDataframe):
                cols = value.columns
                self.__featureEngineering = cols


        else:
            if isinstance(value, pandasDataframe):
                cols = value.columns
                cols = [col for col in cols if col not in self.__dataSourcing]
                # cols = cols.drop(self.dataSourcing)
                self.__featureEngineering = cols

            if isinstance(value, pysparkDataframe):
                cols = value.columns
                cols = [col for col in cols if col not in self.__dataSourcing]
                self.__featureEngineering = cols

        if isinstance(value, pandasDataframe):
            if self.speedUp == False:
                self.__correlation = value.corr(method='pearson')

        if isinstance(value, pysparkDataframe):
            if self.speedUp == False:
                spark = SparkSession.builder.appName("vevesta").getOrCreate()
                columnNames = []
                columnNames = value.columns
                for i in range(len(value.columns)):
                    value = value.withColumn(columnNames[i], value[columnNames[i]].cast(DoubleType()))
                vectorCol = "corrfeatures"
                assembler = VectorAssembler(inputCols=value.columns, outputCol=vectorCol)
                df_vector = assembler.transform(value).select(vectorCol)
                matrix = Correlation.corr(df_vector, vectorCol).collect()[0][0]
                corrMatrix = matrix.toArray().tolist()
                dfCorr = spark.createDataFrame(corrMatrix, columnNames)
                self.__correlation = dfCorr

    @property
    def fe(self):
        return self.__featureEngineering

    @featureEngineering.setter
    def fe(self, value):
        self.featureEngineering = value

    def startModelling(self):
        self.__startlocals = dict(getmembers(stack()[1][0]))['f_locals'].copy()

    def endModelling(self):

        temp = dict(getmembers(stack()[1][0]))['f_locals'].copy()
        self.temp = getmembers(stack()[1])

        self.__variables = {**self.__variables, **{i: temp.get(i) for i in temp if
                                                   i not in self.__startlocals and i[0] != '_' and (
                                                           type(temp[i]) in self.__primitiveDataTypes or isinstance(
                                                       temp[i], (str, int, float, bool)))}}

        return self.__variables

    # create alias of method modellingStart and modellingEnd
    start = startModelling
    end = endModelling
    @property
    def Y(self):
        return self.__Y
            
    @Y.setter
    def Y(self,value):
        if isinstance(value,Series):
            if value.size==self.__sampleSize:
                self.__Y=value
                self.__v=None
            else:
                print('Panda series size not matching with the dataframe size')
        elif isinstance(value,str):
            if value in self.__data.columns:
                self.__Y=self.__data[value]
                self.__v=value
            else:
                print("Column not found")

    # function to get arguments of a function
    def param(self, **decoratorparam):
        def params(functionName):
            def wrapper(*args, **kwargs):
                # to get parameters of function that are passed
                functionParameters = signature(functionName).bind(*args, **kwargs).arguments
                functionParameters = dict(functionParameters)
                # to get values that are not passed and defautlt values
                defaultParameters = signature(functionName)
                for param in defaultParameters.parameters.values():
                    # checks in key exist in abouve dictionary then doesn't update it will default value, otherwise append into dictionary
                    if (param.default is not param.empty) and (param.name not in functionParameters):
                        functionParameters[param.name] = param.default

                for key, value in decoratorparam.items():
                    if (key in functionParameters) and (type[value] in self.__primitiveDataTypes):
                        functionParameters[value] = functionParameters.pop(key)

                self.__variables = {**self.__variables, **{key: value for key, value in functionParameters.items() if
                                                           type(value) in [int, float, bool,
                                                                           str] and key not in self.__variables}}

            return wrapper

        return params

    # Exp = Experiment
    # -------------
    def __getMessage(self):
        messagesList = [
            "For additional features, explore our tool at https://www.vevesta.com?utm_source=vevestaX for free.",
            "Track evolution of Data Science projects at https://www.vevesta.com?utm_source=vevestaX for free.",
            "Manage notes, codes and models in one single place by using our tool at https://www.vevesta.com?utm_source=vevestaX",
            "For faster discovery of features, explore our tool at https://www.vevesta.com?utm_source=vevestaX",
            "Find the right technique for your Machine Learning project at https://www.vevesta.com?utm_source=vevestaX",
            "Give us a reason to celebrate, give us your feedback at vevestax@vevesta.com",
            "Give us a reason to cheer, give us a star on Github: https://github.com/Vevesta/VevestaX",
            "Mail us at vevestax@vevesta.com to follow the latest updates to the library, VevestaX.",
            "Help your ML community work better by giving us a Github star at https://github.com/Vevesta/VevestaX",
            "Easily organize and manage your notes, documents, code, data and models. Explore Vevesta at https://www.vevesta.com?utm_source=vevestaX",
            "Spread the word in your ML community by giving us a Github star at https://github.com/Vevesta/VevestaX",
            "Love VevestaX? Give us a shoutout at vevestax@vevesta.com",
            "Get access to latest release ahead of others by subscribing to vevestax@vevesta.com"
        ]
        return (messagesList[randint(0, len(messagesList) - 1)])

    def __colorCellExcel(self, val):
        if -1 <= val <= -0.9:
            color = '#0D47A1'
            return 'background-color: %s' % color
        elif -0.9 < val <= -0.8:
            color = '#1565C0'
            return 'background-color: %s' % color
        elif -0.8 < val <= -0.7:
            color = '#1976D2'
            return 'background-color: %s' % color
        elif -0.7 < val <= -0.6:
            color = '#2962FF'
            return 'background-color: %s' % color
        elif -0.6 < val <= -0.5:
            color = '#2979FF'
            return 'background-color: %s' % color
        elif -0.5 < val <= -0.4:
            color = '#1E88E5'
            return 'background-color: %s' % color
        elif -0.4 < val <= -0.3:
            color = '#42A5F5'
            return 'background-color: %s' % color
        elif -0.3 < val <= -0.2:
            color = '#64B5F6'
            return 'background-color: %s' % color
        elif -0.2 < val <= -0.1:
            color = '#90CAF9'
            return 'background-color: %s' % color
        elif -0.1 < val <= 0:
            color = '#BBDEFB'
            return 'background-color: %s' % color
        elif val == 0:
            color = '#E3F2FD'
            return 'background-color: %s' % color
        elif 0 < val <= 0.1:
            color = '#F1F8E9'
            return 'background-color: %s' % color
        elif 0.1 < val <= 0.2:
            color = '#DCEDC8'
            return 'background-color: %s' % color
        elif 0.2 < val <= 0.3:
            color = '#C5E1A5'
            return 'background-color: %s' % color
        elif 0.3 < val <= 0.4:
            color = '#AED581'
            return 'background-color: %s' % color
        elif 0.4 < val <= 0.5:
            color = '#9CCC65'
            return 'background-color: %s' % color
        elif 0.5 < val <= 0.6:
            color = '#7CB342'
            return 'background-color: %s' % color
        elif 0.6 < val <= 0.7:
            color = '#689F38'
            return 'background-color: %s' % color
        elif 0.7 < val <= 0.8:
            color = '#558B2F'
            return 'background-color: %s' % color
        elif 0.8 < val <= 0.9:
            color = '#33691E'
            return 'background-color: %s' % color
        elif 0.9 < val <= 1:
            color = '#2E7D32'
            return 'background-color: %s' % color

    def __textColor(self, val):
        if -0.1 < val < 0.1:
            color = 'black'
        else:
            color = 'white'
        return 'color: %s' % color

    def __profilingReport(self, fileName):
        sheetName = 'Profiling Report'
        if self.speedUp:
            return
        if not isinstance(self.__data, pandasDataframe):
            return
        if self.__data.empty or len(self.__data) == 0:
            return
        if fileName is None:
            return print("Error: Provide the Excel File")

        data = [
            {'Number_of_observation': self.__data.shape[0],
             'Number_of_variables': self.__data.shape[1],
             'Missing_cells': self.__data.isna().sum().sum(),
             'Missing_cells(%)': (self.__data.isnull().sum().sum() * 100) / (
                     self.__data.notnull().sum().sum() + self.__data.isnull().sum().sum()),
             'Duplicate_rows': self.__data.duplicated().sum(),
             'Duplicate_rows(%)': (self.__data.duplicated().sum() * 100) / len(self.__data),
             'Total_size_in_memory(byte)': self.__data.memory_usage().sum(),
             'Average_record_size_in_memory(byte)': self.__data.memory_usage().sum() / len(self.__data)
             }
        ]
        profilingDataframe = pandasDataframe(data)

        profileOfVariableDataframe = pandasDataframe(
            {"Field Name": ["Distinct", "Distinct (%)", "Missing", "Missing (%)", "Infinite", "Infinite (%)", "Mean",
                            "Minimum", "Maximum", "Zeros", "Zeros (%)", "Negative", "Negative (%)", "Kurtosis",
                            "Skewness", "Median", "Mode", "Outliers", "Outliers (%)", "Q1 quantile", "Q2 quantile",
                            "Q3 quantile", "100th quantile", "Total Memory Size(bytes)"]})

        numericColumns = self.__data.select_dtypes(include=["number"]).columns
        for col in numericColumns:
            # finding outliers for each column
            Q1 = quantile(self.__data[col], 0.25)
            Q3 = quantile(self.__data[col], 0.75)
            IQR = Q3 - Q1
            outlier = ((self.__data[col] < (Q1 - 1.5 * IQR)) | (self.__data[col] > (Q3 + 1.5 * IQR))).sum()
            col_dict = {"Distinct": self.__data[col].nunique(),
                        "Distinct (%)": self.__data[col].nunique() * 100 / self.__data.shape[0],
                        "Missing": self.__data[col].isna().sum(),
                        "Missing (%)": (self.__data[col].isnull().sum() * 100) / (self.__data.shape[0]),
                        "Infinite": isinf(self.__data[col]).values.sum(),
                        "Infinite (%)": isinf(self.__data[col]).values.sum() * 100 / (self.__data.shape[0]),
                        "Mean": self.__data[col].mean(),
                        "Minimum": self.__data[col].min(),
                        "Maximum": self.__data[col].max(),
                        "Zeros": (self.__data[col] == 0).sum(),
                        "Zeros (%)": (self.__data[col] == 0).sum() * 100 / self.__data.shape[0],
                        "Negative": (self.__data[col] < 0).sum(),
                        "Negative (%)": (self.__data[col] < 0).sum() * 100 / self.__data.shape[0],
                        "Kurtosis": kurtosis(self.__data[col], axis=0, bias=True),
                        "Skewness": skew(self.__data[col], axis=0, bias=True),
                        "Median": median(self.__data[col]),
                        "Mode": mode(self.__data[col]),
                        "Outliers": outlier,
                        "Outliers (%)": outlier * 100 / self.__data.shape[0],
                        "Q1 quantile": quantile(self.__data[col], 0.25),
                        "Q2 quantile": quantile(self.__data[col], 0.5),
                        "Q3 quantile": quantile(self.__data[col], 0.75),
                        "100th quantile": quantile(self.__data[col], 1),
                        "Total Memory Size(bytes)": self.__data[col].memory_usage()}
            profileOfVariableDataframe[col] = col_dict.values()

        nonNumericalColumns = self.__data.select_dtypes(exclude=["number", "datetime"]).columns
        for col in nonNumericalColumns:
            col_dict = {"Distinct": self.__data[col].nunique(),
                        "Distinct (%)": self.__data[col].nunique() * 100 / self.__data.shape[0],
                        "Missing": self.__data[col].isna().sum(),
                        "Missing (%)": (self.__data[col].isnull().sum() * 100) / (self.__data.shape[0]),
                        "Infinite": "NA",
                        "Infinite (%)": "NA",
                        "Mean": "NA",
                        "Minimum": "NA",
                        "Maximum": "NA",
                        "Zeros": "NA",
                        "Zeros (%)": "NA",
                        "Negative": "NA",
                        "Negative (%)": "NA",
                        "Kurtosis": "NA",
                        "Skewness": "NA",
                        "Median": "NA",
                        "Mode": "NA",
                        "Outliers": "NA",
                        "Outliers (%)": "NA",
                        "Q1 quantile": "NA",
                        "Q2 quantile": "NA",
                        "Q3 quantile": "NA",
                        "100th quantile": "NA",
                        "Total Memory Size(bytes)": self.__data[col].memory_usage()}
            profileOfVariableDataframe[col] = col_dict.values()

        if isfile(fileName):
            with ExcelWriter(fileName, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                profilingDataframe.to_excel(writer, sheet_name="Profiling Report", index=False)
                profileOfVariableDataframe.to_excel(writer, sheet_name="Variables Data Profile", index=False)

    def dump(self, techniqueUsed, filename=None, message=None, version=None, showMessage=True, repoName=None):

        existingData = None
        modelingData = None
        featureEngineeringData = None
        messageData = None
        experimentID = 1
        localTimestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        mode = 'w'

        if (filename == None):
            filename = "vevesta.xlsx"

        # updating variables
        # when no V.start & v.end are not called, all variables in the code get tracked or in colab/kaggle where all variables will get tracked
        if (len(self.__variables) == 0):
            temp = dict(getmembers(stack()[1][0]))['f_locals'].copy()
            self.__variables = {**{i: temp.get(i) for i in temp if i[0] != '_' and (
                    type(temp[i]) in self.__primitiveDataTypes or isinstance(temp[i], (str, int, float, bool)))}}

        # check if file already exists
        if (isfile(filename)):
            # mode = 'a'
            existingData = read_excel(filename, sheet_name='dataSourcing', index_col=[])

            featureEngineeringData = read_excel(filename, sheet_name='featureEngineering', index_col=[])

            modelingData = read_excel(filename, sheet_name='modelling', index_col=[])

            messageData = read_excel(filename, sheet_name='messages', index_col=[])

            experimentID = max(modelingData["experimentID"]) + 1

        if self.__dataSourcing is None:
            df_dataSourcing = pandasDataframe(index=[1])

        else:
            df_dataSourcing = pandasDataframe(1, index=[1], columns=self.__dataSourcing)

        df_dataSourcing.insert(0, 'experimentID', experimentID)
        df_dataSourcing = concat([existingData, df_dataSourcing], ignore_index=True).fillna(0)

        if self.featureEngineering is None:
            df_featureEngineering = pandasDataframe(index=[1])

        else:
            df_featureEngineering = pandasDataframe(1, index=[1], columns=self.featureEngineering)

        df_featureEngineering.insert(0, 'experimentID', experimentID)
        df_featureEngineering = concat([featureEngineeringData, df_featureEngineering],
                                              ignore_index=True).fillna(0)

        if self.__dataSourcing is None and self.__featureEngineering is None:
            modeling = pandasDataframe(
                data={**{'experimentID': experimentID, 'timestamp': localTimestamp},
                      **{k: [v] for k, v in self.__variables.items()}}, index=[0])
        elif self.__dataSourcing is None:
            modeling = pandasDataframe(data={
                **{'experimentID': experimentID, 'features': ','.join(self.featureEngineering),
                   'timestamp': localTimestamp}, **{k: [v] for k, v in self.__variables.items()}},
                index=[0])
        elif self.__featureEngineering is None:
            modeling = pandasDataframe(data={**{'experimentID': experimentID, 'features': ','.join(self.dataSourcing),
                                                 'timestamp': localTimestamp},
                                              **{k: [v] for k, v in self.__variables.items()}}, index=[0])
        else:
            modeling = pandasDataframe(data={**{'experimentID': experimentID,
                                                 'features': ','.join(self.dataSourcing) + ',' + ','.join(
                                                     self.featureEngineering),
                                                 'timestamp': localTimestamp},
                                              **{k: [v] for k, v in self.__variables.items()}}, index=[0])

        modeling = concat([modelingData, modeling], ignore_index=True)

        # message table
        data = {
            'experimentID': experimentID,
            'techniqueUsed': techniqueUsed,
            'message': message,
            'version': version,
            'filename': self.__filename,
            'timestamp': localTimestamp
        }

        df_messages = pandasDataframe(index=[1], data=data)
        df_messages = concat([messageData, df_messages], ignore_index=True)

        self.__sampleSize = 100 if self.__sampleSize >= 100 else self.__sampleSize

        if isinstance(self.__data, pandasDataframe):
            sampledData = self.__data.sample(self.__sampleSize)

        if isinstance(self.__data, pysparkDataframe):
            if self.__data.count() >= 100:
                sampledData = self.__data.sample(100 / self.__data.count())

            if self.__data.count() < 100:
                sampledData = self.__data.sample(1.0)

        with ExcelWriter(filename, engine='openpyxl') as writer:

            df_dataSourcing.to_excel(writer, sheet_name='dataSourcing', index=False)

            df_featureEngineering.to_excel(writer, sheet_name='featureEngineering', index=False)
            modeling.to_excel(writer, sheet_name='modelling', index=False)

            df_messages.to_excel(writer, sheet_name='messages', index=False)

            if isinstance(self.__data, pandasDataframe):
                pandasDataframe(sampledData).to_excel(writer, sheet_name='sampledata', index=False)

            if isinstance(self.__data, pysparkDataframe):
                sampledData.toPandas().to_excel(writer, sheet_name='sampledata', index=False)

            if self.speedUp == False:
                if self.__correlation is not None:
                    if isinstance(sampledData, pandasDataframe):
                        pandasDataframe(self.__correlation).style. \
                            applymap(self.__colorCellExcel). \
                            applymap(self.__textColor). \
                            to_excel(writer, sheet_name='EDA-correlation', index=True)

                    if isinstance(sampledData, pysparkDataframe):
                        correlation = self.__correlation.toPandas()
                        correlation.set_index(correlation.columns, inplace=True)
                        pandasDataframe(correlation).style. \
                            applymap(self.__colorCellExcel). \
                            applymap(self.__textColor). \
                            to_excel(writer, sheet_name='EDA-correlation', index=True)

        self.__profilingReport(filename)

        if self.speedUp == False:
            self.__EDA(filename)

        self.__plot(filename)

        print("Dumped the experiment in the file " + filename)

        if showMessage:
            console_message = self.__getMessage()
            print(console_message)

        # log to tool
        backend_url = 'https://api.matrixkanban.com/services-1.0-SNAPSHOT'
        access_token = self.__find_access_token()
        payload = {
            'macAddress': str(getnode()),
            'accessToken': access_token,
            'vevestaXVersion': __version__,
            'platform': {
                'bits': platform.architecture()[0],
                'linkage': platform.architecture()[1],
                'machine': platform.machine(),
                'system_platform': platform.platform(),
                'processor': platform.processor(),
                'python_build_no': platform.python_build()[0],
                'python_build_date': platform.python_build()[1],
                'python_compiler': platform.python_compiler(),
                'python_branch': platform.python_branch(),
                'python_implementation': platform.python_implementation(),
                'python_revision': platform.python_revision(),
                'python_version': platform.python_version(),
                'system_release': platform.release(),
                'os_name': platform.system(),
                'system_release_version': platform.version(),
            }
        }
        headers_for_log = {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*'
        }
        requests.post(url=backend_url + '/DumpLog', headers=headers_for_log, data=json.dumps(payload))

        # push to git
        if repoName is not None:
            try:
                git_token = self.__find_git_token(is_v_commit=False)
                self.__git_commit(git_token=git_token, repo_name=repoName, branch_name=techniqueUsed,
                                  commitMessage=message)
                print('File pushed to git')
            except Exception as e:
                print('File not pushed to git')

    def __EDA(self, fileName):
        if isinstance(self.__data, pandasDataframe):
            self.__EDAForPandas(fileName)

    def __EDAForPandas(self, fileName):

        if self.__data.empty or len(self.__data) == 0:
            return

        if not isinstance(self.__data, pandasDataframe):
            return

        if (fileName == None):
            return print("Error: Provide the Excel File to plot the models")

        columnTextImgone = 'B2'
        columnTextImgtwo = 'B44'
        directoryToDumpData = 'vevestaXDump'
        # creating a new folder in current directory
        Path(directoryToDumpData).mkdir(parents=True, exist_ok=True)
        ValueImageFile = "Value.jpeg"
        ValueRatioImageFile = "ValuePerFeature.jpeg"
        NumericalFeatureDistributionImageFile = "NumericalFeatureDistribution.jpeg"
        NonNumericFeaturesImgFile = "NonNumericFeatures.jpeg"
        FeatureHistogramImageFile = "FeatureHistogram.jpeg"
        OutliersImageFile = "Outliers.jpeg"
        #NumericFeatures3Dplots = "NumericFeatures3Dplots.jpeg"
        ProbabilityDensityFunction="ProbabilityDensityFunction.jpeg"

        # EDA missing values
        plt.figure(figsize=(6, 6))
        plt.imshow(self.__data.isna(), aspect="auto", interpolation="nearest", cmap="coolwarm", extent=[0, 7, 0, 7])
        plt.title("Sample Number vs Column Number")
        plt.xlabel("Column Number")
        plt.ylabel("Sample Number")
        plt.savefig(join(directoryToDumpData, ValueImageFile), bbox_inches='tight', dpi=100)
        plt.close()

        # eda numeric feature distribution
        RatioData = self.__data.isna().mean().sort_values()
        xAxis = list(RatioData.index)
        yAxis = list(RatioData)
        plt.figure(figsize=(6, 6))
        plt.bar(xAxis, yAxis)
        plt.title("Percentage of missing values per feature")
        plt.xlabel("Feature Names")
        plt.ylabel("Ratio of missing values per feature")
        plt.xticks(rotation=90)
        plt.savefig(join(directoryToDumpData, ValueRatioImageFile), bbox_inches='tight', dpi=100)
        plt.close()

        # eda non numeric feature distribution
        self.__data.plot(lw=0, marker="x", subplots=True, layout=(-1, 4), figsize=(10, 10), markersize=5,
                         title="Numeric feature Distribution(with X-axis representing the position in the file)").flatten()
        plt.tight_layout()
        plt.savefig(join(directoryToDumpData, NumericalFeatureDistributionImageFile), bbox_inches='tight',
                    dpi=100)
        plt.close()

        # EDA for outliers
        numericColumns = self.__data.select_dtypes(include=["number"])
        red_circle = dict(markerfacecolor='red', marker='o', markeredgecolor='white')
        fig, axs = plt.subplots(2, len(numericColumns.columns)//2, figsize=(10, 10))
        fig.suptitle('Outliers',fontsize=20)
        for i, ax in enumerate(axs.flat):
            ax.boxplot(numericColumns.iloc[:, i], flierprops=red_circle)
            ax.set_title(self.__data.columns[i], fontsize=15)
            #ax.tick_params(axis='both', labelrotation=45)
            plt.subplots_adjust(wspace=2)
            plt.savefig(join(directoryToDumpData, OutliersImageFile), bbox_inches='tight', dpi=100)
        plt.close()

        # EDA for 3D-Plots
        """
        numericDataframe = self.__data.select_dtypes(include='number')
        fig = plt.figure(figsize=(200, 80))
        k = 1
        for pair in itertools.combinations(numericDataframe.columns, 3):
            if k > 100:
                break
            ax = fig.add_subplot(len(numericDataframe.columns), len(numericDataframe.columns), k, projection='3d')
            ax.scatter3D(numericDataframe[pair[0]], numericDataframe[pair[1]], numericDataframe[pair[2]])
            ax.set_xlabel(pair[0])
            ax.set_ylabel(pair[1])
            ax.set_zlabel(pair[2])
            plt.subplots_adjust(wspace=1)
            k += 1
        plt.savefig(os.path.join(directoryToDumpData, NumericFeatures3Dplots), bbox_inches='tight', dpi=100)
        plt.close()"""

        # Identify non-numerical features
        nonNumericalColumns = self.__data.select_dtypes(exclude=["number", "datetime"])
        if len(nonNumericalColumns.columns) != 0:
            fig = plt.figure(figsize=(7, 7))
            k = 1
            for col in nonNumericalColumns.columns:
                ax = fig.add_subplot(len(nonNumericalColumns.columns), len(nonNumericalColumns.columns), k)
                nonNumericalColumns[col].value_counts(sort=True)[0:10].plot(kind='bar', logy=False, title=col, lw=0,
                                                                            ax=ax)
                k += 1
            plt.savefig(join(directoryToDumpData, NonNumericFeaturesImgFile), bbox_inches='tight', dpi=100)
            plt.close()

        # feature distribution
        fig = self.__data.hist(bins=len(self.__data), figsize=(30, 25), layout=(-1, 3), edgecolor="black",
                               xlabelsize=15, ylabelsize=15)
        [x.title.set_size(15) for x in fig.ravel()]
        [x.tick_params(axis='x', labelrotation=90) for x in fig.ravel()]
        plt.plot()
        plt.suptitle('Feature Histogram',fontsize=20)
        plt.savefig(join(directoryToDumpData, FeatureHistogramImageFile), bbox_inches='tight', dpi=100)
        plt.close()
        
        #Probability Density Function
        numericDataframe = self.__data.select_dtypes(include='number')  
        if self.__Y is not None and (self.__Y.dtype=='int64' or self.__Y.dtype=='object') and self.__Y.dtype!='float64':
            k=1
            fig = plt.figure(figsize=(20,15))
            for i in numericDataframe:
                if i!=self.__v:
                    ax = fig.add_subplot(4,(len(numericDataframe.columns)//4)+1, k)
                    frequency=self.__Y.value_counts().keys().tolist()[0:10]
                    y=self.__Y[self.__Y.isin(frequency)]
                    sns.kdeplot(x=numericDataframe[i],hue=y, ax = ax,fill=True)
                    k+=1
            plt.savefig(join(directoryToDumpData, ProbabilityDensityFunction), bbox_inches='tight', dpi=100)
            plt.close()
        
            

              
        pdffile="EDA.pdf"
        images=[ValueImageFile,ValueRatioImageFile,NumericalFeatureDistributionImageFile,NonNumericFeaturesImgFile,FeatureHistogramImageFile,OutliersImageFile,ProbabilityDensityFunction]
        a4inputsize = (mm_to_pt(210),mm_to_pt(297))
        layout_function = get_layout_fun(a4inputsize)
        file=[]
        for i in images:
            if exists(join(directoryToDumpData,i)):
                file.append(join(directoryToDumpData,i))
        with open(pdffile,"wb") as f:          
            f.write(convert(file,layout_fun=layout_function))

        

        if (isfile(fileName)):
            workBook = load_workbook(fileName)
            workBook.create_sheet('EDA-missingValues')
            plotSheet = workBook['EDA-missingValues']
            img = Image(join(directoryToDumpData, ValueImageFile))
            img.anchor = columnTextImgone
            plotSheet.add_image(img)
            

            image = Image(join(directoryToDumpData, ValueRatioImageFile))
            image.anchor = columnTextImgtwo
            plotSheet.add_image(image)
             

            # adding the plot for the Numeric Fetaure Distribution
            workBook.create_sheet('EDA-NumericfeatureDistribution')
            fetaureplotsheet = workBook['EDA-NumericfeatureDistribution']
            featureImg = Image(
                join(directoryToDumpData, NumericalFeatureDistributionImageFile))
            featureImg.anchor = columnTextImgone
            fetaureplotsheet.add_image(featureImg)
            


            # adding boxplot for Numeric features
            workBook.create_sheet('EDA-Boxplot')
            outlierplotsheet = workBook['EDA-Boxplot']
            OutlierImg = Image(
                join(directoryToDumpData, OutliersImageFile))
            OutlierImg.anchor = columnTextImgone
            outlierplotsheet.add_image(OutlierImg)
            

            # adding 3D plots for numeric features
            """
            workBook.create_sheet('EDA-3Dplot')
            ThreeDplotsheet = workBook['EDA-3Dplot']
            ThreeDImg = openpyxl.drawing.image.Image(
                os.path.join(directoryToDumpData, NumericFeatures3Dplots))
            ThreeDImg.anchor = columnTextImgone
            ThreeDplotsheet.add_image(ThreeDImg)"""

            # adding non-numeric column
            nonNumericalColumns = self.__data.select_dtypes(exclude=["number", "datetime"])
            if len(nonNumericalColumns.columns) != 0 and exists(
                    join(directoryToDumpData, NonNumericFeaturesImgFile)):
                workBook.create_sheet('EDA-NonNumericFeatures')
                nonNumericPlotSheet = workBook['EDA-NonNumericFeatures']
                nonNumericFeatureImage = Image(
                    join(directoryToDumpData, NonNumericFeaturesImgFile))
                nonNumericFeatureImage.anchor = columnTextImgone
                nonNumericPlotSheet.add_image(nonNumericFeatureImage)
            
            

            if exists(join(directoryToDumpData, FeatureHistogramImageFile)):
                workBookName = 'EDA-Feature Histogram'
                workBook.create_sheet(workBookName)
                featureDistribution = workBook[workBookName]
                featureDistributionImage = Image(
                    join(directoryToDumpData, FeatureHistogramImageFile))
                featureDistributionImage.anchor = columnTextImgone
                featureDistribution.add_image(featureDistributionImage)
                
            if exists(join(directoryToDumpData, ProbabilityDensityFunction)):
                workBookName = 'EDA-PDF'
                workBook.create_sheet(workBookName)
                pdfPlotsheet = workBook[workBookName]
                pdfImage = Image(
                    join(directoryToDumpData, ProbabilityDensityFunction))
                pdfImage.anchor = columnTextImgone
                pdfPlotsheet.add_image(pdfImage)   
            
                         
            workBook.save(fileName)
        workBook.close()

    def __plot(self, fileName):

        modelingData = None

        if (fileName == None):
            return print("Error: Provide the Excel File to plot the models")

        sheetName = 'modelling'
        modelingData = self.__getExcelSheetData(fileName, sheetName)

        # returns nothing if dataframe is empty
        if modelingData.empty:
            return

        # excluding numeric, datetime type
        nonNumericColumns = modelingData.select_dtypes(exclude=['number', 'datetime'])
        modelingData.drop(nonNumericColumns, axis=1, inplace=True)
        modelingData.drop('experimentID', axis=1, inplace=True)

        # checks if there are any columns after the timestamp column
        if len(modelingData.columns) == 0:
            return

        directoryToDumpData = 'vevestaXDump'
        self.__truncateFolder(directoryToDumpData)
        # creating a new folder in current directory
        Path(directoryToDumpData).mkdir(parents=True, exist_ok=True)

        # checks if file exist then only loads it and create a new sheet for plots
        if (isfile(fileName)):
            workBook = load_workbook(fileName)
            workBook.create_sheet('performancePlots')
            plotSheet = workBook['performancePlots']
            xAxis = list(nonNumericColumns['timestamp'])
            columnValue = 2

            for column in modelingData.columns:
                yAxis = list(modelingData[column])

                imageName = str(column) + '.jpeg'
                columnText = 'B'
                columnText += str(columnValue)

                # creates a seperate plots for every timestamp vs column and saves it
                fig, ax = plt.subplots()
                ax.plot(xAxis, yAxis, linestyle='-', marker='o')
                # size of plot horizontally fixed to 5 inches and height to 5 inches
                plt.gcf().set_size_inches(13, 5)
                # rotating the x axis labels
                plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right', fontsize='x-small')

                plt.title('Timestamp vs ' + str(column))
                plt.xlabel('Timestamp')
                plt.ylabel(str(column))

                plt.savefig(join(directoryToDumpData, imageName), bbox_inches='tight', dpi=100)
                plt.close()

                img = Image(join(directoryToDumpData, imageName))
                img.anchor = columnText
                plotSheet.add_image(img)

                columnValue += 27

            workBook.save(fileName)
        workBook.close()

    # to truncate the content inside the vevestaXDump folder if it exist
    def __truncateFolder(self, folderName):
        if isdir(folderName):
            for filename in listdir(folderName):
                file_path = join(folderName, filename)
                try:
                    if isfile(file_path) or islink(file_path):
                        unlink(file_path)
                    elif isdir(file_path):
                        rmtree(file_path)
                except Exception as e:
                    print('Failed to delete %s. Reason: %s' % (file_path, e))

    # generic function to get modelling sheet data from any excel file with sheetName modelling
    def __getExcelSheetData(self, fileName, sheetName):
        # checks if file exist then only if fetches modelling data
        if (isfile(fileName)):
            excelFile = load_workbook(fileName, read_only=True)
            # check if modelling sheet exist in excel file
            if sheetName in excelFile.sheetnames:
                modelingData = read_excel(fileName, sheet_name=sheetName, index_col=[])
                return modelingData

    def commit(self, techniqueUsed, filename=None, message=None, version=None, projectId=None,
               repoName=None, branch=None):
        self.dump(techniqueUsed, filename=filename, message=message, version=version, showMessage=False, repoName=None)

        # api-endpoint
        token = self.__find_access_token()
        backend_url = 'https://api.matrixkanban.com/services-1.0-SNAPSHOT'

        # push to git
        try:
            git_token = self.__find_git_token(is_v_commit=True, backend_url=backend_url, access_token=token)
            if repoName is None:
                project = self.__fetch_project(backend_url=backend_url, access_token=token, projectId=projectId)
                if project['gitRepoName'] != '':
                    repoName = project['gitRepoName']
                else:
                    repoName = project['title']

            if branch is None:
                branch = techniqueUsed
            self.__git_commit(git_token=git_token, repo_name=repoName, branch_name=branch, commitMessage=message)
            print('File pushed to git')
        except Exception as e:
            print('File not pushed to git')

        # upload attachment
        filename = self.get_filename()
        file_exists = exists(filename)
        if file_exists:
            files = {'file': open(filename, 'rb')}
            headers_for_file = {'Authorization': 'Bearer ' + token}
            params = {'taskId': 0}
            response = requests.post(url=backend_url + '/Attachments', headers=headers_for_file, params=params,
                                         files=files)
            attachments = list()
            attachments.append(response.json())
            files = {'file': open('EDA.pdf', 'rb')}
            response = requests.post(url=backend_url + '/Attachments', headers=headers_for_file, params=params,
                                         files=files)
            attachments.append(response.json())

        # upload note
        headers_for_note = {
            'Authorization': 'Bearer ' + token,
            'Access-Control-Allow-Origin': '*',
            'Accept': '*/*',
            'Content-Type': 'application/json'
        }
        payload = {
            "projectId": projectId,
            "title": techniqueUsed,
            "message": message,
            "modeling": self.__variables,
            "dataSourced": self.__dataSourcing,
            "featureEngineered": self.__featureEngineering
        }
        
        if file_exists:
            payload['attachments'] = attachments
        else:
            payload['errorMessage'] = 'File not pushed to Vevesta'
            print('File not pushed to Vevesta')
        response = requests.post(url=backend_url + '/VevestaX', headers=headers_for_note, data=json.dumps(payload))
        if response.status_code == 200:
            print("Wrote experiment to tool, Vevesta")
        else:
            print("Failed to write experiment to tool, Vevesta")

    def __git_commit(self, git_token, repo_name, branch_name='main', commitMessage=None):
        g = Github(git_token)

        # format repository name and branch name according to GitHub naming conventions
        repo_name = sub(r'[^A-Za-z0-9_.-]', '-', repo_name)
        branch_name = sub(r'[^A-Za-z0-9_.\-/]', '_', branch_name)

        # find the repo or create a new repo if not exist
        user = g.get_user()
        repo = self.__find_repo(user, repo_name)
        if repo is None:
            repo = user.create_repo(repo_name,
                                    has_issues=True,
                                    has_wiki=True,
                                    has_downloads=True,
                                    has_projects=True,
                                    auto_init=True,
                                    allow_squash_merge=True,
                                    allow_merge_commit=True,
                                    allow_rebase_merge=True)

        # check if branch exist or create a new branch (if not exist) from main or master
        try:
            repo.get_branch(branch_name)
        except GithubException:
            try:
                source_branch = repo.get_branch('main')
            except GithubException:
                source_branch = repo.get_branch('master')
            repo.create_git_ref(ref='refs/heads/' + branch_name, sha=source_branch.commit.sha)

        # push file to git
        file_name = self.get_filename()
        if exists(file_name):
            file_content = self.__read_file(file_name)

            # update if file exists, else create a new file
            try:
                contents = repo.get_contents(file_name, ref=branch_name)
                if commitMessage is None:
                    commitMessage = 'updated ' + file_name
                repo.update_file(
                    contents.path,
                    commitMessage,
                    file_content,
                    sha=contents.sha,
                    branch=branch_name
                )
            except GithubException:
                if commitMessage is None:
                    commitMessage = 'added ' + file_name
                repo.create_file(file_name, commitMessage, file_content, branch=branch_name)

    def __find_repo(self, github_user, repo_name):
        all_repos = github_user.get_repos()
        repos = list(filter(lambda r: r.name.casefold() == repo_name.casefold(), all_repos))
        return repos[0] if len(repos) > 0 else None
